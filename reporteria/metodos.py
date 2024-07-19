import openpyxl
import os

carpeta = "media/excel_files"

def agregarReporte(pNombre,openRt,reporte):
    # Crear un nuevo libro de trabajo y una hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MisDatos"

    nombre = pNombre+".xlsx"
    ruta = os.path.join(carpeta, nombre)

    if not os.path.exists(carpeta):
        os.makedirs(carpeta)

    ws['A1'] = 'Nombre'
    ws['A2'] = 'anno_mes_rt'
    ws['A3'] = 'entrega_total'
    ws['A4'] = 'entrega_a_tiempo'
    ws['A5'] = 'stock_productos'
    ws['A6'] = 'stock_proveedores'
    ws['A7'] = 'adquisicion_recibida'
    ws['A8'] = 'venta_pedido'
    ws['A9'] = 'venta_realizada'
    ws['A10'] = 'factura_total'
    ws['A11'] = 'factura_pagada'
    ws['A12'] = 'boleta_total'
    ws['A13'] = 'boleta_pagada'
    
    ws['B1'] = 'Valor'
    ws['B2'] = reporte.anno_mes_rt
    ws['B3'] = reporte.entrega_total
    ws['B4'] = reporte.entrega_a_tiempo
    ws['B5'] = reporte.stock_productos
    ws['B6'] = reporte.pedido_proveedor
    ws['B7'] = reporte.adquisicion_recibida
    ws['B8'] = reporte.venta_pedido
    ws['B9'] = reporte.venta_realizada
    ws['B10'] = reporte.factura_total
    ws['B11'] = reporte.factura_pagada
    ws['B12'] = reporte.boleta_total
    ws['B13'] = reporte.boleta_pagada

    # Guardar el archivo
    wb.save(ruta)
    if openRt:
        abrirReporte(ruta)

def abrirReporte(ruta):
    # Verificar que el archivo se haya guardado correctamente
    if os.path.exists(ruta):
        print(f"El archivo de la ruta '{ruta}' existe")
        # Abrir el archivo con la aplicación predeterminada
    else:
        print(f"El archivo de la ruta '{ruta}' No Existe")
